from pathlib import Path

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill

OUTPUT_DIRECTORY = Path("data/processed")


def format_worksheet(worksheet) -> None:
    """Aplica formatação profissional na planilha."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in worksheet.columns:
        largest_value = max(len(str(cell.value or "")) for cell in column_cells)

        column_letter = column_cells[0].column_letter

        worksheet.column_dimensions[column_letter].width = min(largest_value + 2, 32)


def add_temperature_chart(worksheet) -> None:
    """Adiciona o gráfico de temperaturas diárias."""

    chart = LineChart()

    chart.title = "Previsão diária de temperatura"
    chart.y_axis.title = "Temperatura (°C)"
    chart.x_axis.title = "Data"
    chart.height = 8
    chart.width = 16
    chart.style = 13

    temperature_data = Reference(
        worksheet,
        min_col=3,
        max_col=5,
        min_row=1,
        max_row=worksheet.max_row,
    )

    date_categories = Reference(
        worksheet,
        min_col=1,
        min_row=2,
        max_row=worksheet.max_row,
    )

    chart.add_data(
        temperature_data,
        titles_from_data=True,
        from_rows=False,
    )

    chart.set_categories(date_categories)
    chart.x_axis.number_format = "yyyy-mm-dd"
    chart.legend.position = "r"

    for series, title in zip(
        chart.series,
        ["Mínima", "Média", "Máxima"],
        strict=True,
    ):
        series.tx = SeriesLabel(v=title)

    worksheet.add_chart(chart, "I2")


def export_weather_data(
    hourly_dataframe: pd.DataFrame,
    daily_dataframe: pd.DataFrame,
) -> tuple[Path, Path]:
    """Exporta os dados para CSV e Excel."""

    OUTPUT_DIRECTORY.mkdir(
        parents=True,
        exist_ok=True,
    )

    csv_path = OUTPUT_DIRECTORY / "weather_data.csv"
    excel_path = OUTPUT_DIRECTORY / "weather_report.xlsx"
    daily_excel_dataframe = daily_dataframe.copy()
    daily_excel_dataframe["date"] = daily_excel_dataframe["date"].astype(str)

    hourly_dataframe.to_csv(
        csv_path,
        index=False,
        encoding="utf-8-sig",
    )

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
    ) as writer:
        hourly_dataframe.to_excel(
            writer,
            sheet_name="Hourly Data",
            index=False,
        )

        daily_excel_dataframe.to_excel(
            writer,
            sheet_name="Daily Summary",
            index=False,
        )

        hourly_worksheet = writer.book["Hourly Data"]
        daily_worksheet = writer.book["Daily Summary"]

        format_worksheet(hourly_worksheet)
        format_worksheet(daily_worksheet)

        for cell in hourly_worksheet["A"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"

        for cell in daily_worksheet["A"][1:]:
            cell.number_format = "yyyy-mm-dd"

        add_temperature_chart(daily_worksheet)

    return csv_path, excel_path
