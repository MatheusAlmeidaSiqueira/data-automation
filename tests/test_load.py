from pathlib import Path

from openpyxl import load_workbook

from data_automation import load
from data_automation.transform import create_daily_summary, transform_weather_data
from tests.test_transform import sample_weather_data


def test_export_weather_data_creates_csv_and_excel(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(load, "OUTPUT_DIRECTORY", tmp_path)
    hourly = transform_weather_data(sample_weather_data(), "Guarulhos")
    daily = create_daily_summary(hourly)

    csv_path, excel_path = load.export_weather_data(hourly, daily)

    assert csv_path.exists()
    assert excel_path.exists()

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == ["Hourly Data", "Daily Summary"]
    assert len(workbook["Daily Summary"]._charts) == 1
